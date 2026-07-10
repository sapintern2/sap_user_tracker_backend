from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.classification_movement import ClassificationMovement
from app.models.daily_user import DailyUser
from app.models.deleted_user import DeletedUser
from app.models.upload import Upload
from app.services.comparison import normalize_category


HEADER_FILL = PatternFill("solid", fgColor="DDEBE7")
NOTE_FILL = PatternFill("solid", fgColor="F8E7C8")
TITLE_FONT = Font(bold=True)


def _category_key(category: str | None) -> str:
    key = normalize_category(category)
    if key == "advanced_users":
        return "advanced"
    if key == "core_users":
        return "core"
    if key == "self_service_users":
        return "self"
    return "other"


def _movement_key(from_category: str | None, to_category: str | None) -> str:
    return f"{normalize_category(from_category)}_to_{normalize_category(to_category)}"


def build_master_audit_workbook(db: Session) -> BytesIO:
    uploads = db.scalars(
        select(Upload).order_by(Upload.upload_date.asc(), Upload.id.asc())
    ).all()

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    movement_sheet = workbook.create_sheet("Classification Movements")
    deleted_sheet = workbook.create_sheet("Deleted Users")
    summary_sheet = workbook.create_sheet("Daily Summary")

    _fill_data_sheet(data_sheet, db, uploads)
    _fill_classification_movements_sheet(movement_sheet, db)
    _fill_deleted_users_sheet(deleted_sheet, db)
    _fill_summary_sheet(summary_sheet, db, uploads)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _fill_data_sheet(sheet, db: Session, uploads: list[Upload]) -> None:
    for index, upload in enumerate(uploads):
        start_column = 1 + (index * 3)
        user_column = start_column
        category_column = start_column + 1
        separator_column = start_column + 2

        deleted_count = db.scalar(
            select(func.count(DeletedUser.id)).where(DeletedUser.current_upload_id == upload.id)
        ) or 0

        note = f"{deleted_count} removed" if deleted_count else "No removals"
        sheet.cell(row=2, column=user_column, value=note)
        sheet.cell(row=2, column=user_column).fill = NOTE_FILL
        sheet.cell(row=2, column=user_column).font = TITLE_FONT

        sheet.cell(row=4, column=user_column, value=upload.upload_date)
        sheet.cell(row=4, column=category_column, value=upload.total_users)
        sheet.cell(row=4, column=user_column).number_format = "yyyy-mm-dd"
        sheet.cell(row=4, column=user_column).font = TITLE_FONT
        sheet.cell(row=4, column=category_column).font = TITLE_FONT

        sheet.cell(row=5, column=user_column, value="User")
        sheet.cell(row=5, column=category_column, value="Target Classification")
        for column in (user_column, category_column):
            cell = sheet.cell(row=5, column=column)
            cell.fill = HEADER_FILL
            cell.font = TITLE_FONT

        users = db.scalars(
            select(DailyUser)
            .where(DailyUser.upload_id == upload.id)
            .order_by(DailyUser.username)
        ).all()
        for row_index, user in enumerate(users, start=6):
            sheet.cell(row=row_index, column=user_column, value=user.username)
            sheet.cell(row=row_index, column=category_column, value=user.category)

        sheet.column_dimensions[get_column_letter(user_column)].width = 22
        sheet.column_dimensions[get_column_letter(category_column)].width = 28
        sheet.column_dimensions[get_column_letter(separator_column)].width = 4

    sheet.freeze_panes = "A6"


def _fill_classification_movements_sheet(sheet, db: Session) -> None:
    headers = [
        "Movement Date",
        "Username",
        "From Classification",
        "To Classification",
        "Previous Upload Date",
        "Current Upload Date",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = TITLE_FONT

    movements = db.scalars(
        select(ClassificationMovement).order_by(
            ClassificationMovement.movement_date,
            ClassificationMovement.username,
        )
    ).all()
    for movement in movements:
        sheet.append(
            [
                movement.movement_date,
                movement.username,
                movement.from_category,
                movement.to_category,
                movement.previous_upload.upload_date,
                movement.current_upload.upload_date,
            ]
        )

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=6):
        row[0].number_format = "yyyy-mm-dd"
        row[4].number_format = "yyyy-mm-dd"
        row[5].number_format = "yyyy-mm-dd"

    widths = [18, 22, 28, 28, 22, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"


def _fill_deleted_users_sheet(sheet, db: Session) -> None:
    headers = ["Deleted Date", "Username", "Target Classification", "Last Seen Date"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = TITLE_FONT

    deleted_users = db.scalars(
        select(DeletedUser).order_by(DeletedUser.deleted_date, DeletedUser.username)
    ).all()
    for user in deleted_users:
        sheet.append([user.deleted_date, user.username, user.category, user.last_seen_date])

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=4):
        row[0].number_format = "yyyy-mm-dd"
        row[3].number_format = "yyyy-mm-dd"

    widths = [18, 22, 28, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"


def _fill_summary_sheet(sheet, db: Session, uploads: list[Upload]) -> None:
    headers = [
        "Date",
        "Total Users",
        "Advanced Users",
        "Core Users",
        "Self-Service Users",
        "Other Users",
        "Deleted Users",
        "Advanced to Core",
        "Advanced to Self-Service",
        "Core to Advanced",
        "Core to Self-Service",
        "Self-Service to Advanced",
        "Self-Service to Core",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = TITLE_FONT

    for upload in uploads:
        users = db.scalars(select(DailyUser).where(DailyUser.upload_id == upload.id)).all()
        counts = {"advanced": 0, "core": 0, "self": 0, "other": 0}
        for user in users:
            counts[_category_key(user.category)] += 1

        deleted_count = len(
            db.scalars(
                select(DeletedUser).where(DeletedUser.current_upload_id == upload.id)
            ).all()
        )
        movement_counts = {
            "advanced_users_to_core_users": 0,
            "advanced_users_to_self_service_users": 0,
            "core_users_to_advanced_users": 0,
            "core_users_to_self_service_users": 0,
            "self_service_users_to_advanced_users": 0,
            "self_service_users_to_core_users": 0,
        }
        movements = db.scalars(
            select(ClassificationMovement).where(
                ClassificationMovement.current_upload_id == upload.id
            )
        ).all()
        for movement in movements:
            key = _movement_key(movement.from_category, movement.to_category)
            if key in movement_counts:
                movement_counts[key] += 1

        sheet.append(
            [
                upload.upload_date,
                upload.total_users,
                counts["advanced"],
                counts["core"],
                counts["self"],
                counts["other"],
                deleted_count,
                movement_counts["advanced_users_to_core_users"],
                movement_counts["advanced_users_to_self_service_users"],
                movement_counts["core_users_to_advanced_users"],
                movement_counts["core_users_to_self_service_users"],
                movement_counts["self_service_users_to_advanced_users"],
                movement_counts["self_service_users_to_core_users"],
            ]
        )

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        row[0].number_format = "yyyy-mm-dd"

    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 20
        for cell in sheet[get_column_letter(column)]:
            cell.alignment = Alignment(horizontal="left")

    sheet.freeze_panes = "A2"
